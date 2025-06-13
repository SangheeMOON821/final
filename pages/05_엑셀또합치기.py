import streamlit as st
from io import BytesIO
import re
from openpyxl import load_workbook, Workbook

st.title("엑셀 파일 → 시트 병합기 (스타일 보존)")

uploaded_files = st.file_uploader(
    "엑셀 파일을 여러 개 업로드하세요", 
    type=["xlsx", "xls"],
    accept_multiple_files=True
)

def sanitize_sheet_name(name):
    # Excel 시트명은 최대 31자, 특수문자 불가
    name = re.sub(r'[:\\/?*\[\]]', '', name)
    return name[:31]

if uploaded_files:
    output = BytesIO()
    # 새 워크북 생성 후 기본 시트는 삭제
    target_wb = Workbook()
    target_wb.remove(target_wb.active)

    for uploaded_file in uploaded_files:
        try:
            # 원본 파일 로드 (스타일 정보 포함)
            source_wb = load_workbook(uploaded_file, data_only=False)
            # 첫 번째 시트만 복제하고 싶다면 아래처럼; 전체 시트를 복제하려면 for sheet in source_wb.worksheets 로 변경
            source_sheet = source_wb[source_wb.sheetnames[0]]  
            # 파일 이름을 기반으로 시트명 생성
            base_name = uploaded_file.name.rsplit('.', 1)[0]
            sheet_name = sanitize_sheet_name(base_name)
            ws = target_wb.create_sheet(title=sheet_name)

            # 열 너비 복사
            for col, dim in source_sheet.column_dimensions.items():
                ws.column_dimensions[col].width = dim.width

            # 행 높이 복사
            for row, dim in source_sheet.row_dimensions.items():
                if dim.height:
                    ws.row_dimensions[row].height = dim.height

            # 병합 셀 복사
            for merged in source_sheet.merged_cells.ranges:
                ws.merge_cells(str(merged))

            # Freeze panes 복사
            ws.freeze_panes = source_sheet.freeze_panes

            # 각 셀 복사 (값 + 스타일)
            for row in source_sheet.iter_rows():
                for cell in row:
                    new_cell = ws.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        new_cell.font = cell.font
                        new_cell.border = cell.border
                        new_cell.fill = cell.fill
                        new_cell.number_format = cell.number_format
                        new_cell.protection = cell.protection
                        new_cell.alignment = cell.alignment

        except Exception as e:
            st.error(f"{uploaded_file.name} 읽기 실패: {e}")

    # 저장
    target_wb.save(output)
    output.seek(0)

    st.success(f"{len(uploaded_files)}개의 파일이 스타일까지 보존하여 병합되었습니다!")

    st.download_button(
        label="엑셀로 다운로드",
        data=output.getvalue(),
        file_name="시트별_스타일_보존_병합.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
