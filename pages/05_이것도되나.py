import streamlit as st
from io import BytesIO
import re
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from copy import copy

st.title("엑셀 파일 → 시트 병합기 (완전 스타일 & 크기 보존)")

uploaded_files = st.file_uploader(
    "엑셀 파일을 여러 개 업로드하세요", 
    type=["xlsx", "xls"],
    accept_multiple_files=True
)

def sanitize_sheet_name(name):
    # Excel 시트명은 최대 31자, 특수문자 불가
    name = re.sub(r'[:\\/?*\[\]]', '', name)
    return name[:31]

if uploaded_files:
    output = BytesIO()
    # 첫번째 파일명(확장자 제외)
    first_base = uploaded_files[0].name.rsplit('.', 1)[0]
    count = len(uploaded_files)
    # 새 워크북 생성, 기본 시트 제거
    target_wb = Workbook()
    target_wb.remove(target_wb.active)

    for uploaded_file in uploaded_files:
        try:
            # 스타일 포함 로드
            src_wb = load_workbook(uploaded_file, data_only=False)
            src = src_wb[src_wb.sheetnames[0]]
            title = sanitize_sheet_name(uploaded_file.name.rsplit('.', 1)[0])
            tgt = target_wb.create_sheet(title=title)

            # 기본 행 높이/열 너비 복제
            default_row_h = src.sheet_format.defaultRowHeight
            default_col_w = src.sheet_format.defaultColWidth
            tgt.sheet_format.defaultRowHeight = default_row_h
            tgt.sheet_format.defaultColWidth = default_col_w

            max_row = src.max_row
            max_col = src.max_column

            # 모든 열 너비·속성 복사
            for col_idx in range(1, max_col+1):
                col = get_column_letter(col_idx)
                src_dim = src.column_dimensions.get(col, None)
                tgt_dim = tgt.column_dimensions[col]
                if src_dim and src_dim.width is not None:
                    tgt_dim.width = src_dim.width
                else:
                    tgt_dim.width = default_col_w
                tgt_dim.hidden = src_dim.hidden if src_dim else False
                tgt_dim.outlineLevel = src_dim.outlineLevel if src_dim else 0
                tgt_dim.bestFit = src_dim.bestFit if src_dim else False

            # 모든 행 높이·속성 복사
            for row_idx in range(1, max_row+1):
                src_dim = src.row_dimensions.get(row_idx, None)
                tgt_dim = tgt.row_dimensions[row_idx]
                if src_dim and src_dim.height is not None:
                    tgt_dim.height = src_dim.height
                else:
                    tgt_dim.height = default_row_h
                tgt_dim.hidden = src_dim.hidden if src_dim else False
                tgt_dim.outlineLevel = src_dim.outlineLevel if src_dim else 0

            # 병합 셀 복사
            for merged in src.merged_cells.ranges:
                tgt.merge_cells(str(merged))

            # Freeze panes 복사
            tgt.freeze_panes = src.freeze_panes

            # 셀 값 + 스타일 복사
            for row in src.iter_rows():
                for cell in row:
                    nc = tgt.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        nc.font = copy(cell.font)
                        nc.border = copy(cell.border)
                        nc.fill = copy(cell.fill)
                        nc.number_format = cell.number_format
                        nc.protection = copy(cell.protection)
                        nc.alignment = copy(cell.alignment)

        except Exception as e:
            st.error(f"{uploaded_file.name} 읽기 실패: {e}")

    # 저장
    target_wb.save(output)
    output.seek(0)

    filename = f"{first_base} {count} (병합).xlsx"
    st.success(f"{count}개의 파일이 완전 보존되어 병합되었습니다!")
    st.download_button(
        label="엑셀로 다운로드",
        data=output.getvalue(),
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
